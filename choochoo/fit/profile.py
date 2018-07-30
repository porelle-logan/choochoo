
import datetime as dt
from abc import abstractmethod
from os.path import dirname, join
from pickle import dump, load
from re import compile
from struct import unpack

import openpyxl as xls
from more_itertools import peekable
from pkg_resources import resource_stream

from ..args import PATH
from ..log import make_log

LITTLE, BIG = 0, 1
PROFILE = 'global-profile.pkl'
HEADER_GLOBAL_TYPE = -1
HEADER_FIELDS = [
    ('header_size', 1, 'uint8'),
    ('protocol_version', 1, 'uint8'),
    ('profile_version', 1, 'uint16'),
    ('data_size', 1, 'uint32'),
    ('fit_text', 4, 'string'),
    ('checksum', 1, 'uint16')
]


def package_fit_profile(args):
    log = make_log(args)
    in_path = args.file(PATH, 0)
    log.info('Reading from %s' % in_path)
    nlog, types, messages = read_profile(log, in_path)
    out_path = join(dirname(__file__), PROFILE)
    nlog.set_log(None)
    log.info('Writing to %s' % out_path)
    with open(out_path, 'wb') as output:
        dump((nlog, types, messages), output)
    # test loading
    log.info('Test loading from %r' % PROFILE)
    log.info('Loaded %s, %s' % load_profile(log))


def read_profile(log, path):
    nlog = NullableLog(log)
    wb = xls.load_workbook(path)
    types = Types(nlog, wb['Types'])
    messages = Messages(nlog, wb['Messages'], types)
    return nlog, types, messages


def load_profile(log):
    input = resource_stream(__name__, PROFILE)
    nlog, types, messages = load(input)
    nlog.set_log(log)
    return types, messages


class NullableLog:

    def __init__(self, log):
        self.set_log(log)

    def set_log(self, log):
        self.__log = log

    def debug(self, *args):
        self.__log.debug(*args)

    def info(self, *args):
        self.__log.info(*args)

    def warn(self, *args):
        self.__log.warn(*args)

    def error(self, *args):
        self.__log.error(*args)


class Named:
    """
    Has a name.  Base for both fields and messages
    """

    def __init__(self, log, name):
        self._log = log
        self.name = name

    def __str__(self):
        return '%s: %s' % (self.__class__.__name__, self.name)


class ErrorDict(dict):

    def __init__(self, log, error_msg):
        self.__log = log
        self.__error_msg = error_msg
        super().__init__()

    def add_named(self, item):
        self[item.name] = item

    def __getitem__(self, item):
        try:
            return super().__getitem__(item)
        except KeyError:
            msg = self.__error_msg % item
            self.__log.error(msg)
            raise KeyError(msg)


class ErrorList(list):

    def __init__(self, log, error_msg):
        self.__log = log
        self.__error_msg = error_msg
        super().__init__()

    def __getitem__(self, item):
        try:
            return super().__getitem__(item)
        except IndexError:
            msg = self.__error_msg % item
            self.__log.error(msg)
            raise IndexError(msg)


class AbstractType(Named):

    def __init__(self, log, name, size, base_type=None):
        super().__init__(log, name)
        self.base_type = base_type
        self.size = size

    @abstractmethod
    def profile_to_internal(self, cell_contents):
        raise NotImplementedError('%s: %s' % (self.__class__.__name__, self.name))

    @abstractmethod
    def raw_to_internal(self, bytes, count, endian):
        raise NotImplementedError('%s: %s' % (self.__class__.__name__, self.name))


class BaseType(AbstractType):

    def __init__(self, log, name, size, func):
        super().__init__(log, name, size)
        self.__func = func

    def profile_to_internal(self, cell_contents):
        return self.__func(cell_contents)


class StructSupport(BaseType):

    def _pack_bad(self, value):
        bad = (bytearray(self.size), bytearray(self.size))
        for endian in (LITTLE, BIG):
            bytes = value
            for i in range(self.size):
                j = i if endian == LITTLE else self.size - i - 1
                bad[endian][j] = bytes & 0xff
                bytes >>= 8
        return bad

    def _is_bad(self, data, bad):
        size = len(bad)
        count = len(data) // size
        return all(bad == data[size*i:size*(i+1)] for i in range(count))

    def _unpack(self, data, formats, bad, count, endian):
        if self._is_bad(data, bad[endian]):
            return None
        else:
            value = unpack(formats[endian] % count, data[0:count * self.size])
            if count == 1:
                value = value[0]
            else:
                value = list(value)
            return value


class String(BaseType):

    def __init__(self, log, name):
        super().__init__(log, name, 1, str)

    def raw_to_internal(self, bytes, count, endian):
        return str(b''.join(unpack('%dc' % count, bytes)), encoding='utf-8')


class Boolean(BaseType):

    def __init__(self, log, name):
        super().__init__(log, name, 1, bool)

    def raw_to_internal(self, bytes, count, endian):
        bools = [bool(byte) for byte in bytes]
        if count == 1:
            return bools[0]
        else:
            return bools


class AutoInteger(StructSupport):

    pattern = compile(r'^([su]?)int(\d{1,2})(z?)$')

    size_to_format = {1: 'b', 2: 'h', 4: 'i', 8: 'q'}

    def __init__(self, log, name):
        match = self.pattern.match(name)
        self.signed = match.group(1) != 'u'
        bits = int(match.group(2))
        if bits % 8:
            raise Exception('Size of %r not a multiple of 8 bits' % name)
        super().__init__(log, name, bits // 8, self.int)
        if self.size not in self.size_to_format:
            raise Exception('Cannot unpack %d bytes as an integer' % self.size)
        format = self.size_to_format[self.size]
        if not self.signed:
            format = format.upper()
        self.formats = ['<%d' + format, '>%d' + format]
        self.bad = self._pack_bad(0 if match.group(3) == 'z' else 2 ** (bits - (1 if self.signed else 0)) - 1)

    @staticmethod
    def int(cell):
        if isinstance(cell, int):
            return cell
        else:
            return int(cell, 0)

    def raw_to_internal(self, data, count, endian):
        return self._unpack(data, self.formats, self.bad, count, endian)


class AliasInteger(AutoInteger):

    def __init__(self, log, name, spec):
        super().__init__(log, spec)
        self.name = name


class Date(AliasInteger):

    def __init__(self, log, name, utc):
        super().__init__(log, name, 'uint32')
        self.__tzinfo = dt.timezone.utc if utc else None

    def raw_to_internal(self, data, count, endian):
        time = super().raw_to_internal(data, count, endian)
        if time >= 0x10000000:
            time = dt.datetime(1989, 12, 31, tzinfo=self.__tzinfo) + dt.timedelta(seconds=time)
        return time


class AutoFloat(StructSupport):

    pattern = compile(r'^float(\d{1,2})$')

    size_to_format = {2: 'e', 4: 'f', 8: 'd'}

    def __init__(self, log, name):
        match = self.pattern.match(name)
        bits = int(match.group(1))
        if bits % 8:
            raise Exception('Size of %r not a multiple of 8 bits' % name)
        super().__init__(log, name, bits // 8, float)
        if self.size not in self.size_to_format:
            raise Exception('Cannot unpack %d bytes as a float' % self.size)
        format = self.size_to_format[self.size]
        self.formats = ['<%d' + format, '>%d' + format]
        self.bad = self._pack_bad(2 ** bits - 1)

    def raw_to_internal(self, data, count, endian):
        return self._unpack(data, self.formats, self.bad, count, endian)


class Mapping(AbstractType):

    def __init__(self, log, row, rows, types):
        name = row[0]
        base_type_name = row[1]
        base_type = types.profile_to_type(base_type_name, auto_create=True)
        super().__init__(log, name, base_type.size, base_type=base_type)
        self._profile_to_internal = ErrorDict(log, 'No internal value for profile %r')
        self._internal_to_profile = ErrorDict(log, 'No profile value for internal %r')
        for row in rows:
            if row[0] or row[2] is None or row[3] is None:
                rows.prepend(row)
                break
            self.__add_mapping(row)
        log.debug('Parsed %d values' % len(self._profile_to_internal))

    def profile_to_internal(self, cell_contents):
        return self._profile_to_internal[cell_contents]

    def internal_to_profile(self, value):
        return self._internal_to_profile[value]

    def raw_to_internal(self, bytes, size, endian):
        return self.base_type.raw_to_internal(bytes, size, endian)

    def __add_mapping(self, row):
        profile = row[2]
        internal = self.base_type.profile_to_internal(row[3])
        self._profile_to_internal[profile] = internal
        self._internal_to_profile[internal] = profile


# table 4-6 of FIT defn doc
BASE_TYPE_NAMES = ['enum', 'sint8', 'uint8', 'sint16', 'uint16', 'sint32', 'uint32',
                   'string', 'float32', 'float64',
                   'uint8z', 'uint16z', 'uint32z', 'byte', 'sint64', 'uint64', 'uint64z']


class Types:

    def __init__(self, log, sheet):
        self.__log = log
        self.__profile_to_type = ErrorDict(log, 'No type for profile %r')
        # these are not 'base types' in the same sense as types having base types.
        # rather, they are the 'base (integer) types' described in the docs
        self.base_types = ErrorList(log, 'No base type for number %r')
        self.__add_known_types()
        rows = peekable([cell.value for cell in row] for row in sheet.iter_rows())
        for row in rows:
            if row[0] and row[0][0].isupper():
                self.__log.debug('Skipping %s' % row)
            elif row[0]:
                self.__log.info('Parsing type %s' % row[0])
                self.__add_type(Mapping(self.__log, row, rows, self))

    def __add_known_types(self):
        # these cannot be inferred from name
        self.__add_type(String(self.__log, 'string'))
        self.__add_type(AliasInteger(self.__log, 'enum', 'uint8'))
        self.__add_type(AliasInteger(self.__log, 'byte', 'uint8'))
        # these can be inferred
        for name in BASE_TYPE_NAMES:
            self.profile_to_type(name, auto_create=True)
            self.base_types.append(self.profile_to_type(name))
        # this is in the spreadsheet, but not in the doc
        self.__add_type(Boolean(self.__log, 'bool'))
        # these are defined in the spreadsheet, but the interpretation is in comments
        self.__add_type(Date(self.__log, 'date_time', True))
        self.__add_type(Date(self.__log, 'local_date_time', False))

    def __add_type(self, type):
        if type.name in self.__profile_to_type:
            duplicate = self.__profile_to_type[type.name]
            if duplicate.size == type.size:
                self.__log.warn('Ignoring duplicate type for %r' % type.name)
            else:
                raise Exception('Duplicate type for %r with differing size (%d  %d)' %
                                (type.name, type.size, duplicate.size))
        else:
            self.__profile_to_type.add_named(type)

    def profile_to_type(self, name, auto_create=False):
        try:
            return self.__profile_to_type[name]
        except KeyError:
            if auto_create:
                for cls in (AutoFloat, AutoInteger):
                    match = cls.pattern.match(name)
                    if match:
                        self.__log.warn('Auto-adding type %s for %r' % (cls.__name__, name))
                        self.__add_type(cls(self.__log, name))
                        return self.profile_to_type(name)
            raise


class MessageField(Named):

    def __init__(self, log, name, number, units, type):
        super().__init__(log, name)
        self.number = number
        self.units = units if units else ''
        self.is_dynamic = self.number is None
        self.type = type

    def profile_to_internal(self, name):
        return self.type.profile_to_internal(name)

    def _with_unit(self, value):
        if value is None:
            return value
        else:
            return str(value) + self.units

    def raw_to_internal(self, data, size, endian, dynamic_cb):
        if self.is_dynamic:
            if not dynamic_cb:
                raise NotImplementedError('Dynamic field (and no callback)')
            for pair in dynamic_cb(self.references):
                try:
                    return self[self.dynamic[pair]].raw_to_internal(data, size, endian, None)
                except KeyError:
                    pass
            raise Exception('No match for dynamic field %r' % self)
        # have to return name because of dynamic fields
        return self.name, self._with_unit(self.type.raw_to_internal(data, size, endian))


class RowMessageField(MessageField):

    def __init__(self, log, row, types):
        super().__init__(log, row[2],
                         int(row[1]) if row[1] is not None else None,
                         row[8],
                         types.profile_to_type(row[3], auto_create=True))


class DynamicMessageField(RowMessageField):

    def __init__(self, log, row, rows, types):
        super().__init__(log, row, types)
        self.__dynamic_tmp_data = []
        self.__dynamic_lookup = ErrorDict(log, 'No dynamic field for %r')
        self.references = set()
        try:
            peek = rows.peek()
            while peek[2] and peek[1] is None:
                row = next(rows)
                for name, value in zip(row[11].split(','), row[12].split(',')):
                    self.__save_dynamic(name.strip(), value.strip(), row)
                peek = rows.peek()
        except StopIteration:
            return

    def __save_dynamic(self, reference_name, reference_value, row):
        self.is_dynamic = True
        self.__dynamic_tmp_data.append((reference_name, reference_value, row))

    def _complete_dynamic(self, message, types):
        for reference_name, reference_value, row in self.__dynamic_tmp_data:
            reference = message.profile_to_field(reference_name)
            value = reference.profile_to_internal(reference_value)
            self.references.add(reference)
            self.__dynamic_lookup[(reference_name, value)] = RowMessageField(self._log, row, types)

    @property
    def dynamic(self):
        return self.__dynamic_lookup


class Message(Named):

    def __init__(self, log, name, number=None):
        super().__init__(log, name)
        if number is not None:
            self.number = number
        self._profile_to_field = ErrorDict(log, 'No field for profile %r')
        self._number_to_field = ErrorDict(log, 'No field for number %r')

    def _add_field(self, field):
        self._profile_to_field.add_named(field)
        self._number_to_field[field.number] = field

    def profile_to_field(self, name):
        return self._profile_to_field[name]

    def number_to_field(self, value):
        return self._number_to_field[value]

    def raw_to_internal(self, data, definition, dynamic_cb=None):
        offset = 0
        message = {}
        for field_desc in definition.fields:
            bytes = data[offset:offset+field_desc.size]
            try:
                field = self.number_to_field(field_desc.number)
                count = field_desc.size // field.type.size
                name, value = self._parse_field(message, field, bytes, count, definition.endian, dynamic_cb)
            except KeyError:
                name = str(field_desc.number)
                count = field_desc.size // field_desc.base_type.size
                value = str(field_desc.base_type.raw_to_internal(bytes, count, definition.endian))
            message[name] = value
            offset += field_desc.size
        return message

    def _parse_field(self, _message, field, bytes, count, endian, dynamic_cb):
        # allow interception for optional field in header
        return field.raw_to_internal(bytes, count, endian, dynamic_cb)


class NumberedMessage(Message):

     def __init__(self, log, name, types):
        try:
            number = types.profile_to_type('mesg_num').profile_to_internal(name)
        except KeyError:
            number = None
            log.warn('No mesg_num for %r' % name)
        super().__init__(log, name, number)


class RowMessage(NumberedMessage):

    def __init__(self, log, row, rows, types):
        super().__init__(log, row[0], types)
        for row in rows:
            if not row[2]:
                rows.prepend(row)
                break
            self.__parse_row(row, rows, types)
        self.__complete_dynamic(types)

    def __parse_row(self, row, rows, types):
        self._add_field(DynamicMessageField(self._log, row, rows, types))

    def __complete_dynamic(self, types):
        # these may be forward references
        for data in self._profile_to_field.values():
            if data.is_dynamic:
                data._complete_dynamic(self, types)


class Header(Message):

    def __init__(self, log, types):
        super().__init__(log, 'HEADER', number=HEADER_GLOBAL_TYPE)
        for n, (name, size, base_type) in enumerate(HEADER_FIELDS):
            self._add_field(MessageField(log, name, n, None, types.profile_to_type(base_type)))

    def _parse_field(self, message, field, data, count, endian, dynamic_cb):
        if field.name == 'checksum' and message['header_size'] == 12:
            return None, None
        else:
            return super()._parse_field(message, field, data, count, endian, dynamic_cb)


class Missing(Message):

    def __init__(self, log, number):
        super().__init__(log, 'MESSAGE %d' % number, number)


class Messages:

    def __init__(self, log, sheet, types):
        self.__log = log
        self.__profile_to_message = ErrorDict(log, 'No message for profile %r')
        self.__number_to_message = ErrorDict(log, 'No message for number %r')
        rows = peekable([cell.value for cell in row] for row in sheet.iter_rows())
        for row in rows:
            if row[0] and row[0][0].isupper():
                self.__log.debug('Skipping %s' % row)
            elif row[0]:
                self.__log.info('Parsing message %s' % row[0])
                self.__add_message(RowMessage(self.__log, row, rows, types))
        self.__add_message(Header(self.__log, types))

    def __add_message(self, message):
        self.__profile_to_message.add_named(message)
        try:
            self.__number_to_message[message.number] = message
        except AttributeError:
            pass

    def profile_to_message(self, name):
        return self.__profile_to_message[name]

    def number_to_message(self, number):
        try:
            return self.__number_to_message[number]
        except KeyError:
            message = Missing(self.__log, number)
            self.__number_to_message[number] = message
            return message
